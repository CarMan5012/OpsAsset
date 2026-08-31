# -*- coding: utf-8 -*-
import os
import sys
import io
import pytest
import pandas as pd
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

TEST_DB_FILE = "./data/pytest_isolated.db"
TEST_DB_URL = f"sqlite:///{TEST_DB_FILE}"
test_engine = create_engine(TEST_DB_URL, connect_args={"check_same_thread": False})
TestSessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=test_engine)

from fastapi.testclient import TestClient
from app.main import app
from app.database import Base, get_db
from app.models import Host, Cluster, HostClusterRelation
from app.routes.config import get_config_from_db
from app.services.excel_service import (
    COL_HOSTNAME, COL_PRIVATE_IP, COL_PUBLIC_IP, COL_CPU, COL_MEMORY,
    COL_DISK, COL_ARCH, COL_OS, COL_KERNEL, COL_ENV, COL_STATUS, COL_CLUSTERS, COL_NOTES
)

def override_get_db():
    db = TestSessionLocal()
    try:
        yield db
    finally:
        db.close()

client = TestClient(app)

def setup_module(module):
    # 覆盖数据库依赖为隔离的测试专用库
    app.dependency_overrides[get_db] = override_get_db
    Base.metadata.create_all(bind=test_engine)
    db = TestSessionLocal()
    get_config_from_db(db)
    db.close()

def teardown_module(module):
    app.dependency_overrides.clear()
    Base.metadata.drop_all(bind=test_engine)
    if os.path.exists(TEST_DB_FILE):
        try:
            os.remove(TEST_DB_FILE)
        except Exception:
            pass

def test_health_check():
    response = client.get("/api/health")
    assert response.status_code == 200
    assert response.json()["status"] == "ok"

def test_meta_config():
    response = client.get("/api/config/meta")
    assert response.status_code == 200
    data = response.json()
    assert "environments" in data
    assert "statuses" in data
    assert "architectures" in data
    assert "cluster_types" in data
    assert "cluster_roles" in data
    assert "os_suggestions" in data
    assert len(data["cluster_types"]) >= 5
    assert "K8s" in data["cluster_roles"]

def test_template_download():
    response = client.get("/api/assets/template")
    assert response.status_code == 200
    assert response.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    df = pd.read_excel(io.BytesIO(response.content))
    assert COL_HOSTNAME in df.columns
    assert COL_PRIVATE_IP in df.columns
    assert COL_CPU in df.columns
    assert COL_ARCH in df.columns

def test_batch_import_and_upsert():
    data = {
        COL_HOSTNAME: ["node-master-01", "node-worker-01", "test-mysql-01"],
        COL_PRIVATE_IP: ["10.0.0.1", "10.0.0.2", "10.0.0.3"],
        COL_PUBLIC_IP: ["1.1.1.1", "", ""],
        COL_CPU: [16, 32, 8],
        COL_MEMORY: [64, 128, 32],
        COL_DISK: [500, 1000, 300],
        COL_ARCH: ["amd64", "amd64", "arm64"],
        COL_OS: ["Ubuntu 22.04", "Ubuntu 22.04", "CentOS 7.9"],
        COL_KERNEL: ["5.15.0", "5.15.0", "3.10.0"],
        COL_ENV: ["生产", "prod", "测试"],
        COL_STATUS: ["在线", "online", "online"],
        COL_CLUSTERS: ["k8s-prod:Master", "k8s-prod:Worker", "mysql-test:Master"],
        COL_NOTES: ["测试节点1", "测试节点2", "数据库测试"]
    }
    df = pd.DataFrame(data)
    excel_stream = io.BytesIO()
    df.to_excel(excel_stream, index=False)
    excel_stream.seek(0)

    # 1. 执行导入
    response = client.post(
        "/api/assets/import?overwrite=true",
        files={"file": ("test_import.xlsx", excel_stream.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    )
    assert response.status_code == 200
    res_json = response.json()
    assert res_json["inserted_count"] == 3
    assert res_json["error_count"] == 0

    # 2. 验证看板概览统计
    dash_res = client.get("/api/dashboard/overview")
    assert dash_res.status_code == 200
    d = dash_res.json()
    assert d["total_hosts"] == 3
    assert d["total_cpu_cores"] == 16 + 32 + 8  # 56
    assert d["total_memory_gb"] == 64 + 128 + 32  # 224
    assert d["prod"]["host_count"] == 2
    assert d["test"]["host_count"] == 1

    # 3. 验证集群自动关联
    clusters_res = client.get("/api/clusters")
    assert clusters_res.status_code == 200
    c_list = clusters_res.json()
    assert len(c_list) == 2
    k8s_c = next(c for c in c_list if c["name"] == "k8s-prod")
    assert k8s_c["node_count"] == 2

    # 4. 测试覆盖更新 (Upsert)
    data_update = {
        COL_HOSTNAME: ["node-master-01-renamed"],
        COL_PRIVATE_IP: ["10.0.0.1"],
        COL_PUBLIC_IP: ["1.1.1.2"],
        COL_CPU: [32],
        COL_MEMORY: [128],
        COL_DISK: [800],
        COL_ARCH: ["amd64"],
        COL_OS: ["Ubuntu 22.04"],
        COL_KERNEL: ["5.15.0-custom"],
        COL_ENV: ["生产"],
        COL_STATUS: ["在线"],
        COL_CLUSTERS: ["k8s-prod:Master"],
        COL_NOTES: ["已更新"]
    }
    df_update = pd.DataFrame(data_update)
    stream_up = io.BytesIO()
    df_update.to_excel(stream_up, index=False)
    stream_up.seek(0)

    res_up = client.post(
        "/api/assets/import?overwrite=true",
        files={"file": ("test_update.xlsx", stream_up.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    )
    assert res_up.status_code == 200
    up_json = res_up.json()
    assert up_json["updated_count"] == 1

    host_list_res = client.get("/api/hosts?keyword=10.0.0.1")
    assert host_list_res.status_code == 200
    h_items = host_list_res.json()["items"]
    assert len(h_items) == 1
    assert h_items[0]["hostname"] == "node-master-01-renamed"
    assert h_items[0]["cpu_cores"] == 32
    assert h_items[0]["memory_gb"] == 128.0
    assert h_items[0]["arch"] == "amd64"

def test_export_assets():
    res_excel = client.get("/api/assets/export?format=xlsx")
    assert res_excel.status_code == 200
    df_out = pd.read_excel(io.BytesIO(res_excel.content))
    assert len(df_out) == 3
    assert "架构" in df_out.columns or "CPU架构" in df_out.columns

    res_csv = client.get("/api/assets/export?format=csv")
    assert res_csv.status_code == 200
    assert "text/csv" in res_csv.headers["content-type"]

def test_update_host_api():
    # 查找已有主机
    h_res = client.get("/api/hosts?keyword=10.0.0.1")
    assert h_res.status_code == 200
    h = h_res.json()["items"][0]
    h_id = h["id"]

    # 执行 PUT 编辑更新
    put_res = client.put(f"/api/hosts/{h_id}", json={
        "hostname": "node-master-01-edited",
        "private_ip": "10.0.0.1",
        "cpu_cores": 64,
        "memory_gb": 256.0,
        "disk_gb": 1000.0,
        "arch": "amd64",
        "os": "Ubuntu 22.04 LTS",
        "kernel_version": "5.15.0-edit",
        "env": "prod",
        "status": "online",
        "notes": "手动编辑测试"
    })
    assert put_res.status_code == 200
    updated_h = put_res.json()
    assert updated_h["hostname"] == "node-master-01-edited"
    assert updated_h["cpu_cores"] == 64
    assert updated_h["updated_at"] is not None

    # 7. 测试 100% 精准查询逻辑
    # 精准检索完整主机名，精确命中 1 条
    exact_res = client.get("/api/hosts?keyword=node-master-01-edited")
    assert exact_res.status_code == 200
    assert exact_res.json()["total"] == 1
    assert exact_res.json()["items"][0]["hostname"] == "node-master-01-edited"

    # 检索不匹配的关键词时返回 0 条
    exact_none = client.get("/api/hosts?keyword=non-existent-host-xyz")
    assert exact_none.status_code == 200
    assert exact_none.json()["total"] == 0

def test_ports_and_range():
    # 1. 验证元数据中的默认端口字典
    meta_res = client.get("/api/config/meta")
    assert meta_res.status_code == 200
    meta = meta_res.json()
    assert "default_middleware_ports" in meta
    assert meta["default_middleware_ports"]["MySQL"] == "3306"
    assert meta["default_middleware_ports"]["Redis"] == "6379"
    assert "30000-32767" in meta["default_middleware_ports"]["K8s"]

    # 2. 创建带端口范围的主机
    host_res = client.post("/api/hosts", json={
        "hostname": "port-test-node",
        "private_ip": "192.168.99.88",
        "public_ip": "1.2.3.4",
        "open_ports": "22, 80, 443, 30000-32767",
        "cpu_cores": 16,
        "memory_gb": 64.0,
        "disk_gb": 500.0,
        "arch": "amd64",
        "os": "CentOS 7.7",
        "kernel_version": "3.10.0-1062.el7.x86_64",
        "env": "prod",
        "status": "online",
        "notes": "端口测试节点"
    })
    assert host_res.status_code == 200
    h = host_res.json()
    assert h["open_ports"] == "22, 80, 443, 30000-32767"

    # 3. 创建带服务端口的集群
    cluster_res = client.post("/api/clusters", json={
        "name": "mysql-port-cluster",
        "cluster_type": "MySQL",
        "port": "3306, 33060",
        "version": "8.0.32",
        "env": "prod",
        "description": "MySQL 双端口实例"
    })
    assert cluster_res.status_code == 200
    c = cluster_res.json()
    assert c["port"] == "3306, 33060"

    # 4. 绑定主机到集群
    bind_res = client.post(f"/api/clusters/{c['id']}/bind-hosts", json={
        "nodes": [{"host_id": h["id"], "role": "Master"}]
    })
    assert bind_res.status_code == 200
    bound_cluster = bind_res.json()
    assert bound_cluster["node_count"] == 1
    assert bound_cluster["nodes"][0]["open_ports"] == "22, 80, 443, 30000-32767"

    # 5. 清理测试数据
    client.delete(f"/api/clusters/{c['id']}")
    client.delete(f"/api/hosts/{h['id']}")

def test_domain_assets_and_dns_check():
    # 1. 创建关联主机
    host_res = client.post("/api/hosts", json={
        "hostname": "domain-gateway-01",
        "private_ip": "192.168.1.200",
        "public_ip": "218.205.231.138",
        "open_ports": "80, 443",
        "cpu_cores": 8,
        "memory_gb": 16.0,
        "disk_gb": 100.0,
        "env": "prod",
        "status": "online",
        "notes": "网关主机"
    })
    assert host_res.status_code == 200
    host_id = host_res.json()["id"]

    # 2. 创建域名资产
    domain_res = client.post("/api/domains", json={
        "domain_name": "https://api.opsasset-test.com/v1",  # 测试自动清理协议与路径
        "public_ip": "218.205.231.138",
        "port": "80, 443",
        "env": "prod",
        "bound_host_id": host_id,
        "notes": "测试核心API域名"
    })
    assert domain_res.status_code == 200
    d = domain_res.json()
    assert d["domain_name"] == "api.opsasset-test.com"
    assert d["public_ip"] == "218.205.231.138"
    assert d["bound_host_id"] == host_id
    domain_id = d["id"]

    # 3. 测试查询列表
    list_res = client.get("/api/domains?keyword=opsasset-test")
    assert list_res.status_code == 200
    assert list_res.json()["total"] == 1
    item = list_res.json()["items"][0]
    assert item["domain_name"] == "api.opsasset-test.com"
    assert item["host"] is not None
    assert item["host"]["hostname"] == "domain-gateway-01"

    # 4. 测试单项 DNS 解析比对接口
    dns_res = client.post(f"/api/domains/{domain_id}/check-dns")
    assert dns_res.status_code == 200
    dns_data = dns_res.json()
    assert "resolve_status" in dns_data
    assert "public_ip" in dns_data

    # 5. 测试修改域名
    put_res = client.put(f"/api/domains/{domain_id}", json={
        "public_ip": "1.2.3.4",
        "notes": "已更新备注"
    })
    assert put_res.status_code == 200
    assert put_res.json()["public_ip"] == "1.2.3.4"
    assert put_res.json()["notes"] == "已更新备注"

    # 6. 测试大盘统计包含了域名数
    dash_res = client.get("/api/dashboard/overview")
    assert dash_res.status_code == 200
    assert dash_res.json()["total_domains"] >= 1

    # 7. 清理测试数据
    del_res = client.delete(f"/api/domains/{domain_id}")
    assert del_res.status_code == 200

    client.delete(f"/api/hosts/{host_id}")


def test_export_custom_columns_and_formatted_data():
    # 1. 创建测试主机
    h1_res = client.post("/api/hosts", json={
        "hostname": "export-test-master-01",
        "private_ip": "172.16.10.10",
        "public_ip": "114.114.114.114",
        "open_ports": "22, 80",
        "cpu_cores": 4,
        "memory_gb": 8.0,
        "disk_gb": 100.0,
        "arch": "amd64",
        "os": "CentOS 7.9",
        "kernel_version": "3.10.0-1160.el7.x86_64",
        "env": "prod",
        "status": "online",
        "notes": "导出测试主机1"
    })
    assert h1_res.status_code == 200
    h1 = h1_res.json()

    h2_res = client.post("/api/hosts", json={
        "hostname": "export-test-worker-02",
        "private_ip": "172.16.10.11",
        "public_ip": "",
        "open_ports": "22",
        "cpu_cores": 8,
        "memory_gb": 2048.0,  # 2 TB
        "disk_gb": 4096.0,    # 4 TB
        "arch": "arm64",
        "os": "Ubuntu 22.04",
        "env": "test",
        "status": "maintenance",
        "notes": "导出测试主机2"
    })
    assert h2_res.status_code == 200
    h2 = h2_res.json()

    # 2. 创建集群并关联到 h1
    c_res = client.post("/api/clusters", json={
        "name": "redis-export-cluster",
        "cluster_type": "Redis",
        "port": "6379",
        "version": "7.0.12",
        "env": "prod",
        "description": "Redis 集群"
    })
    assert c_res.status_code == 200
    c = c_res.json()

    client.post(f"/api/clusters/{c['id']}/bind-hosts", json={
        "nodes": [{"host_id": h1["id"], "role": "Master"}]
    })

    # 3. 测试全量导出并验证格式化数据
    export_all = client.get("/api/assets/export?format=xlsx")
    assert export_all.status_code == 200
    df_all = pd.read_excel(io.BytesIO(export_all.content))
    assert "序号" in df_all.columns
    assert "主机名" in df_all.columns
    assert "环境" in df_all.columns
    assert "状态" in df_all.columns
    assert "CPU" in df_all.columns
    assert "内存" in df_all.columns
    assert "数据盘" in df_all.columns
    assert "所属服务" in df_all.columns
    assert df_all.iloc[0]["序号"] == 1

    # 检查 h1 的数据
    row1 = df_all[df_all["内网IP"] == "172.16.10.10"].iloc[0]
    assert row1["主机名"] == "export-test-master-01"
    assert row1["环境"] == "生产环境"
    assert row1["状态"] == "在线"
    assert row1["CPU"] == "4 核"
    assert row1["内存"] == "8 GB"
    assert row1["内核版本"] == "3.10.0"
    assert "6379" in str(row1["开放端口"])
    assert "Redis" in str(row1["所属服务"])
    assert "v7.0.12" in str(row1["所属服务"])

    # 检查 h2 的 TB 单位换算
    row2 = df_all[df_all["内网IP"] == "172.16.10.11"].iloc[0]
    assert row2["内存"] == "2 TB"
    assert row2["数据盘"] == "4 TB"
    assert row2["环境"] == "测试环境"
    assert row2["状态"] == "维护中"

    # 4. 测试自定义列导出 (包含 index)
    custom_cols = "index,hostname,private_ip,env,status,cpu_cores,memory_gb"
    export_custom = client.get(f"/api/assets/export?format=xlsx&columns={custom_cols}")
    assert export_custom.status_code == 200
    df_custom = pd.read_excel(io.BytesIO(export_custom.content))
    assert list(df_custom.columns) == ["序号", "主机名", "内网IP", "环境", "状态", "CPU", "内存"]
    assert df_custom.iloc[0]["序号"] == 1

    # 5. 测试按指定 ID 导出
    export_ids = client.get(f"/api/assets/export?format=xlsx&ids={h1['id']}")
    assert export_ids.status_code == 200
    df_ids = pd.read_excel(io.BytesIO(export_ids.content))
    assert len(df_ids) == 1
    assert df_ids.iloc[0]["内网IP"] == "172.16.10.10"

    # 6. 测试 CSV 格式导出
    export_csv = client.get(f"/api/assets/export?format=csv&columns={custom_cols}&ids={h1['id']}")
    assert export_csv.status_code == 200
    assert "text/csv" in export_csv.headers["content-type"]
    df_csv = pd.read_csv(io.BytesIO(export_csv.content))
    assert list(df_csv.columns) == ["序号", "主机名", "内网IP", "环境", "状态", "CPU", "内存"]
    assert df_csv.iloc[0]["主机名"] == "export-test-master-01"
    assert df_csv.iloc[0]["序号"] == 1

    # 清理测试数据
    client.delete(f"/api/clusters/{c['id']}")
    client.delete(f"/api/hosts/{h1['id']}")
    client.delete(f"/api/hosts/{h2['id']}")


def test_export_public_ip_merged_cells():
    """测试多台主机共享同一个公网IP时，导出 Excel 会自动跨行合并公网IP单元格"""
    import openpyxl

    shared_pub_ip = "120.55.200.88"
    # 创建2台共享该公网IP的主机
    h1 = client.post("/api/hosts", json={
        "hostname": "pub-merge-node-01",
        "private_ip": "10.0.1.11",
        "public_ip": shared_pub_ip,
        "cpu_cores": 4,
        "memory_gb": 8,
        "disk_gb": 100,
        "env": "prod",
        "status": "online"
    }).json()

    h2 = client.post("/api/hosts", json={
        "hostname": "pub-merge-node-02",
        "private_ip": "10.0.1.12",
        "public_ip": shared_pub_ip,
        "cpu_cores": 4,
        "memory_gb": 8,
        "disk_gb": 100,
        "env": "prod",
        "status": "online"
    }).json()

    try:
        # 指定这两台主机导出
        export_res = client.get(f"/api/assets/export?format=xlsx&ids={h1['id']},{h2['id']}&columns=hostname,private_ip,public_ip")
        assert export_res.status_code == 200

        wb = openpyxl.load_workbook(io.BytesIO(export_res.content))
        ws = wb.active

        # 检查表头
        headers = [cell.value for cell in ws[1]]
        assert "外网IP" in headers
        pub_col_idx = headers.index("外网IP") + 1

        # 检查是否包含合并单元格 (第2行和第3行的第3列)
        merged_ranges = [str(rng) for rng in ws.merged_cells.ranges]
        # 列字母
        col_letter = openpyxl.utils.get_column_letter(pub_col_idx)
        expected_range = f"{col_letter}2:{col_letter}3"
        assert any(expected_range in r for r in merged_ranges), f"未找到预期的公网IP合并单元格 {expected_range}，实际合并范围: {merged_ranges}"

    finally:
        client.delete(f"/api/hosts/{h1['id']}")
        client.delete(f"/api/hosts/{h2['id']}")


def test_excel_styling_features():
    """测试高颜值 Excel 模板与导出文件的样式特性 (表头颜色、冻结窗格、自动筛选、数据验证)"""
    import openpyxl

    # 1. 验证导入模板高颜值特性
    tpl_res = client.get("/api/assets/template")
    assert tpl_res.status_code == 200
    tpl_wb = openpyxl.load_workbook(io.BytesIO(tpl_res.content))
    tpl_ws = tpl_wb.active

    # 验证冻结首行与自动筛选
    assert tpl_ws.freeze_panes == "A2"
    assert tpl_ws.auto_filter.ref is not None
    # 验证表头背景色为 1E293B
    header_fill_color = tpl_ws.cell(row=1, column=1).fill.start_color.rgb
    assert "1E293B" in str(header_fill_color).upper()
    # 验证数据下拉验证已加入
    assert len(tpl_ws.data_validations.dataValidation) >= 1

    # 2. 验证导出 Excel 高颜值特性
    exp_res = client.get("/api/assets/export?format=xlsx")
    assert exp_res.status_code == 200
    exp_wb = openpyxl.load_workbook(io.BytesIO(exp_res.content))
    exp_ws = exp_wb.active

    # 验证冻结首行
    assert exp_ws.freeze_panes == "A2"
    # 验证表头背景色
    exp_header_fill = exp_ws.cell(row=1, column=1).fill.start_color.rgb
    assert "1E293B" in str(exp_header_fill).upper()


def test_database_encryption_mechanism(tmp_path):
    """测试数据库透明加密与存量明文平滑自动迁移机制"""
    from app.database import auto_migrate_plain_to_cipher_if_needed
    import sqlite3

    test_db = tmp_path / "test_asset.db"
    
    # 1. 模拟生成一份标准的 SQLite 明文数据库
    conn = sqlite3.connect(str(test_db))
    cur = conn.cursor()
    cur.execute("CREATE TABLE test_table (id INT, val TEXT);")
    cur.execute("INSERT INTO test_table VALUES (1, 'plain_data');")
    conn.commit()
    conn.close()

    # 验证生成的文件前 16 字节为 SQLite 明文标识
    with open(test_db, "rb") as f:
        header = f.read(16)
    assert header.startswith(b"SQLite format 3")

    # 2. 触发自动平滑转密逻辑
    auto_migrate_plain_to_cipher_if_needed(test_db, "test_secret_key_123456")

    # 3. 验证旧明文库备份文件存在
    bak_file = tmp_path / "test_asset.db.bak_plain"
    assert bak_file.exists()





