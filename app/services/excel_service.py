import io
import re
import ipaddress
from datetime import datetime
import pandas as pd
from typing import Tuple, List, Dict, Any, Optional
from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.models import Host, Cluster, HostClusterRelation
from app.schemas import ImportResultResponse, ImportErrorItem

# 模板列名映射
COL_HOSTNAME = "主机名*"
COL_PRIVATE_IP = "内网IP*"
COL_PUBLIC_IP = "外网IP"
COL_PORTS = "开放端口"
COL_CPU = "CPU(核)*"
COL_MEMORY = "内存(GB)*"
COL_DISK = "数据盘(GB)*"
COL_ARCH = "CPU架构"
COL_OS = "操作系统"
COL_KERNEL = "内核版本"
COL_ENV = "环境*"
COL_STATUS = "状态*"
COL_CLUSTERS = "所属集群/服务(格式: 集群名:角色, 多个用逗号隔开)"
COL_NOTES = "备注"

TEMPLATE_COLUMNS = [
    COL_HOSTNAME,
    COL_PRIVATE_IP,
    COL_PUBLIC_IP,
    COL_PORTS,
    COL_CPU,
    COL_MEMORY,
    COL_DISK,
    COL_ARCH,
    COL_OS,
    COL_KERNEL,
    COL_ENV,
    COL_STATUS,
    COL_CLUSTERS,
    COL_NOTES
]

def is_valid_ipv4(ip_str: str) -> bool:
    if not ip_str or not isinstance(ip_str, str):
        return False
    try:
        ip = ipaddress.IPv4Address(ip_str.strip())
        return True
    except ValueError:
        return False

def normalize_int(val: Any, default: int = 0) -> int:
    if val is None or pd.isna(val):
        return default
    try:
        return int(float(str(val).strip()))
    except (ValueError, TypeError):
        return default

def normalize_float(val: Any, default: float = 0.0) -> float:
    if val is None or pd.isna(val):
        return default
    try:
        return float(str(val).strip())
    except (ValueError, TypeError):
        return default

def normalize_arch(val: Any) -> str:
    if not val:
        return "amd64"
    s = str(val).strip().lower()
    if any(k in s for k in ["amd", "x86", "intel", "x64"]):
        return "amd64"
    if any(k in s for k in ["arm", "aarch", "鲲鹏", "飞腾"]):
        return "arm64"
    if any(k in s for k in ["loong", "龙芯"]):
        return "loongarch64"
    if "mips" in s:
        return "mips64el"
    if "sw" in s or "申威" in s:
        return "sw64"
    return s or "amd64"

def normalize_env(val: Any) -> str:
    if not val:
        return "test"
    s = str(val).strip().lower()
    if s in ["prod", "production", "生产", "生产环境"]:
        return "prod"
    return "test"

def normalize_status(val: Any) -> str:
    if not val:
        return "online"
    s = str(val).strip().lower()
    if s in ["offline", "下线", "离线", "停止", "stopped"]:
        return "offline"
    if s in ["maintenance", "维护", "维护中"]:
        return "maintenance"
    return "online"

def generate_excel_template() -> io.BytesIO:
    """生成标准化 Excel 导入模板 (科技石板深蓝高颜值方案，支持 DataValidation 下拉验证)"""
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = Workbook()
    ws = wb.active
    ws.title = "主机资产导入模板"

    # 1. 科技石板深蓝 (Slate-800) 表头样式 (精致紧凑版)
    header_font = Font(name="Microsoft YaHei", size=10, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=False)
    header_border = Border(
        left=Side(style="thin", color="334155"),
        right=Side(style="thin", color="334155"),
        top=Side(style="thin", color="334155"),
        bottom=Side(style="medium", color="0EA5E9") # 底边青蓝强调线
    )
    thin_border = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1")
    )

    ws.append(TEMPLATE_COLUMNS)
    ws.row_dimensions[1].height = 22

    for col_idx, col_name in enumerate(TEMPLATE_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = header_border

    # 2. 填充 3 条直观高质量的示例数据
    sample_rows = [
        [
            "k8s-prod-master-01",
            "192.168.1.10",
            "114.114.114.10",
            "22, 6443",
            16,
            64,
            500,
            "amd64",
            "Ubuntu 22.04 LTS",
            "5.15.0-89-generic",
            "prod",
            "online",
            "k8s-prod-bj:Master, redis-cluster:Master",
            "核心机房A01机架"
        ],
        [
            "k8s-prod-worker-01",
            "192.168.1.11",
            "",
            "22, 10250",
            32,
            128,
            1000,
            "amd64",
            "Ubuntu 22.04 LTS",
            "5.15.0-89-generic",
            "prod",
            "online",
            "k8s-prod-bj:Worker",
            "核心机房A02机架"
        ],
        [
            "test-db-mysql-01",
            "172.16.10.20",
            "",
            "22, 3306",
            8,
            32,
            300,
            "amd64",
            "CentOS 7.9",
            "3.10.0-1160.el7",
            "test",
            "online",
            "mysql-test-cluster:Master",
            "开发测试数据库"
        ]
    ]

    sample_font = Font(name="Microsoft YaHei", size=9.5, color="1E293B")
    sample_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    center_data_align = Alignment(horizontal="center", vertical="center")
    left_data_align = Alignment(horizontal="left", vertical="center")

    # 环境与状态示例高亮
    prod_fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
    prod_font = Font(name="Microsoft YaHei", size=9.5, color="1D4ED8", bold=True)
    test_fill = PatternFill(start_color="F3E8FF", end_color="F3E8FF", fill_type="solid")
    test_font = Font(name="Microsoft YaHei", size=9.5, color="7E22CE", bold=True)
    online_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
    online_font = Font(name="Microsoft YaHei", size=9.5, color="15803D", bold=True)

    for row_data in sample_rows:
        ws.append(row_data)
        current_row = ws.max_row
        ws.row_dimensions[current_row].height = 20
        for col_idx in range(1, len(row_data) + 1):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.font = sample_font
            cell.border = thin_border
            cell.fill = sample_fill

            # 居中列: 端口、CPU、内存、数据盘、架构、内核、环境、状态
            if col_idx in [4, 5, 6, 7, 8, 10, 11, 12]:
                cell.alignment = center_data_align
            else:
                cell.alignment = left_data_align

            # 特殊色彩标签渲染
            if col_idx == 11: # 环境
                if cell.value == "prod":
                    cell.fill = prod_fill
                    cell.font = prod_font
                elif cell.value == "test":
                    cell.fill = test_fill
                    cell.font = test_font
            elif col_idx == 12: # 状态
                if cell.value == "online":
                    cell.fill = online_fill
                    cell.font = online_font

    # 3. 开启 Excel 数据下拉校验 (Data Validation)
    try:
        # 环境列下拉验证 (K列)
        dv_env = DataValidation(type="list", formula1='"prod,test,dev,stage"', allow_blank=True)
        dv_env.error = '请从下拉列表中选择有效环境 (prod, test, dev, stage)'
        dv_env.errorTitle = '环境输入无效'
        ws.add_data_validation(dv_env)
        dv_env.add("K2:K2000")

        # 状态列下拉验证 (L列)
        dv_status = DataValidation(type="list", formula1='"online,offline,maintenance"', allow_blank=True)
        dv_status.error = '请从下拉列表中选择有效状态 (online, offline, maintenance)'
        dv_status.errorTitle = '状态输入无效'
        ws.add_data_validation(dv_status)
        dv_status.add("L2:L2000")
    except Exception:
        pass

    # 4. 自动设置列宽与视图优化 (紧凑自适应)
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val = str(cell.value or "")
            try:
                val_len = len(val.encode('gbk', 'ignore'))
            except Exception:
                val_len = len(val)
            max_len = max(max_len, val_len)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 10)

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(TEMPLATE_COLUMNS))}{ws.max_row}"
    if ws.views.sheetView:
        ws.views.sheetView[0].showGridLines = True

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def parse_and_import(
    file_bytes: bytes,
    filename: str,
    db: Session,
    overwrite: bool = True
) -> ImportResultResponse:
    """解析上传的 Excel 或 CSV 文件并导入/更新数据库"""
    if filename.endswith(".csv"):
        # 尝试常见编码
        for enc in ["utf-8-sig", "utf-8", "gbk", "gb18030"]:
            try:
                df = pd.read_csv(io.BytesIO(file_bytes), encoding=enc)
                break
            except Exception:
                continue
        else:
            raise ValueError("无法解析 CSV 文件编码，请确保为 UTF-8 或 GBK 格式")
    elif filename.endswith((".xlsx", ".xls")):
        df = pd.read_excel(io.BytesIO(file_bytes))
    else:
        raise ValueError("不支持的文件格式，仅支持 .xlsx, .xls, .csv")

    # 找到关键列的映射（终极健壮容错）
    col_map = {}
    for col in df.columns:
        c_clean = str(col).strip()
        c_lower = c_clean.lower()
        if "主机" in c_clean or "hostname" in c_lower:
            col_map["hostname"] = col
        elif ("内网" in c_clean or "私网" in c_clean or c_lower in ["ip", "private_ip", "private ip"]) and "外网" not in c_clean and "公网" not in c_clean:
            col_map["private_ip"] = col
        elif "外网" in c_clean or "公网" in c_clean or "public_ip" in c_lower or "public ip" in c_lower:
            col_map["public_ip"] = col
        elif "端口" in c_clean or "port" in c_lower or "ports" in c_lower:
            col_map["open_ports"] = col
        elif "架构" in c_clean or "arch" in c_lower:
            col_map["arch"] = col
        elif ("cpu" in c_lower or "算力" in c_clean or ("核" in c_clean and "内核" not in c_clean)) and "架构" not in c_clean and "arch" not in c_lower:
            col_map["cpu_cores"] = col
        elif "内存" in c_clean or "memory" in c_lower or "mem" in c_lower or "ram" in c_lower:
            col_map["memory_gb"] = col
        elif "数据盘" in c_clean or "磁盘" in c_clean or "存储" in c_clean or "disk" in c_lower or "hdd" in c_lower or "ssd" in c_lower:
            col_map["disk_gb"] = col
        elif "内核" in c_clean or "kernel" in c_lower:
            col_map["kernel_version"] = col
        elif "系统" in c_clean or "os" in c_lower:
            col_map["os"] = col
        elif "环境" in c_clean or "env" in c_lower:
            col_map["env"] = col
        elif "状态" in c_clean or "status" in c_lower:
            col_map["status"] = col
        elif "集群" in c_clean or "服务" in c_clean or "cluster" in c_lower:
            col_map["clusters"] = col
        elif "备注" in c_clean or "notes" in c_lower or "说明" in c_clean:
            col_map["notes"] = col

    # 检查必填表头
    required_keys = ["hostname", "private_ip", "cpu_cores", "memory_gb", "disk_gb"]
    missing = [k for k in required_keys if k not in col_map]
    if missing:
        raise ValueError(f"上传文件缺少必要的列: {', '.join(missing)}。请使用标准模板。")

    inserted = 0
    updated = 0
    skipped = 0
    errors: List[ImportErrorItem] = []
    total_rows = len(df)

    def get_val(row, field_key):
        if field_key not in col_map:
            return None
        col_name = col_map[field_key]
        if col_name in row:
            return row[col_name]
        for k in row.index:
            if str(k).strip() == str(col_name).strip():
                return row[k]
        return None

    # 缓存已有的集群与主机避免重复查询
    all_clusters = {c.name: c for c in db.query(Cluster).all()}

    for index, row in df.iterrows():
        row_num = index + 2  # Excel 对应行号（表头为1）
        raw_ip_val = get_val(row, "private_ip")
        raw_ip = str(raw_ip_val).strip() if raw_ip_val is not None else ""

        if not raw_ip or raw_ip.lower() == "nan":
            errors.append(ImportErrorItem(row=row_num, ip=raw_ip or "-", reason="内网IP为空"))
            continue

        if not is_valid_ipv4(raw_ip):
            errors.append(ImportErrorItem(row=row_num, ip=raw_ip, reason=f"内网IP格式不合法: {raw_ip}"))
            continue

        # 解析字段
        try:
            h_name = get_val(row, "hostname")
            hostname = str(h_name).strip() if h_name is not None and str(h_name).lower() != "nan" else f"host-{raw_ip.replace('.', '-')}"

            pub_ip = get_val(row, "public_ip")
            public_ip = str(pub_ip).strip() if pub_ip is not None and str(pub_ip).lower() != "nan" else ""

            ports_raw = get_val(row, "open_ports")
            open_ports = str(ports_raw).strip() if ports_raw is not None and str(ports_raw).lower() != "nan" else ""

            cpu_cores = normalize_int(get_val(row, "cpu_cores"), default=0)
            memory_gb = normalize_float(get_val(row, "memory_gb"), default=0.0)
            disk_gb = normalize_float(get_val(row, "disk_gb"), default=0.0)

            arch_val = normalize_arch(get_val(row, "arch"))

            os_raw = get_val(row, "os")
            os_val = str(os_raw).strip() if os_raw is not None and str(os_raw).lower() != "nan" else ""

            kernel_raw = get_val(row, "kernel_version")
            kernel_val = str(kernel_raw).strip() if kernel_raw is not None and str(kernel_raw).lower() != "nan" else ""

            env_val = normalize_env(get_val(row, "env"))
            status_val = normalize_status(get_val(row, "status"))

            notes_raw = get_val(row, "notes")
            notes_val = str(notes_raw).strip() if notes_raw is not None and str(notes_raw).lower() != "nan" else ""

            clusters_raw = get_val(row, "clusters")
            clusters_str = str(clusters_raw).strip() if clusters_raw is not None and str(clusters_raw).lower() != "nan" else ""

        except Exception as e:
            errors.append(ImportErrorItem(row=row_num, ip=raw_ip, reason=f"数据转换错误: {str(e)}"))
            continue

        # 查询是否已存在
        existing_host = db.query(Host).filter(Host.private_ip == raw_ip).first()

        if existing_host:
            if not overwrite:
                skipped += 1
                continue
            # 覆盖更新
            existing_host.hostname = hostname
            existing_host.public_ip = public_ip
            existing_host.open_ports = open_ports
            existing_host.cpu_cores = cpu_cores
            existing_host.memory_gb = memory_gb
            existing_host.disk_gb = disk_gb
            existing_host.arch = arch_val
            existing_host.os = os_val
            existing_host.kernel_version = kernel_val
            existing_host.env = env_val
            existing_host.status = status_val
            existing_host.notes = notes_val
            existing_host.updated_at = datetime.now()
            target_host = existing_host
            updated += 1
        else:
            new_host = Host(
                hostname=hostname,
                private_ip=raw_ip,
                public_ip=public_ip,
                open_ports=open_ports,
                cpu_cores=cpu_cores,
                memory_gb=memory_gb,
                disk_gb=disk_gb,
                arch=arch_val,
                os=os_val,
                kernel_version=kernel_val,
                env=env_val,
                status=status_val,
                notes=notes_val,
                updated_at=None
            )
            db.add(new_host)
            db.flush()  # 获取 target_host.id
            target_host = new_host
            inserted += 1

        # 处理集群关联
        if clusters_str:
            # 格式: 集群名:角色, 集群2:角色2  或者单纯集群名(默认Worker)
            # 先清除现有对应绑定或者追加
            cluster_entries = [c.strip() for c in re.split(r"[,;，；]", clusters_str) if c.strip()]
            for item in cluster_entries:
                parts = item.split(":") if ":" in item else item.split("：")
                c_name = parts[0].strip()
                c_role = parts[1].strip() if len(parts) > 1 and parts[1].strip() else "Worker"

                if not c_name:
                    continue

                # 查找或新建集群
                if c_name not in all_clusters:
                    # 智能判断集群类型
                    c_type = "Custom"
                    name_lower = c_name.lower()
                    for t in ["k8s", "redis", "mysql", "nacos", "mq", "mongodb", "kafka", "elasticsearch", "clickhouse"]:
                        if t in name_lower:
                            c_type = t.upper() if t in ["k8s", "mq"] else t.capitalize()
                            break

                    new_cluster = Cluster(
                        name=c_name,
                        cluster_type=c_type,
                        env=env_val,
                        description="导入时自动创建"
                    )
                    db.add(new_cluster)
                    db.flush()
                    all_clusters[c_name] = new_cluster

                cluster_obj = all_clusters[c_name]

                # 建立/更新 host_cluster_relation
                rel = db.query(HostClusterRelation).filter(
                    HostClusterRelation.host_id == target_host.id,
                    HostClusterRelation.cluster_id == cluster_obj.id
                ).first()
                if not rel:
                    rel = HostClusterRelation(
                        host_id=target_host.id,
                        cluster_id=cluster_obj.id,
                        role=c_role
                    )
                    db.add(rel)
                else:
                    rel.role = c_role

    db.commit()

    return ImportResultResponse(
        total_rows=total_rows,
        inserted_count=inserted,
        updated_count=updated,
        skipped_count=skipped,
        error_count=len(errors),
        errors=errors
    )


def format_storage_str(gb: Any) -> str:
    """与前端主机资产列表完全一致的存储格式化 (8 GB / 1.5 TB 等，空数据返回空字符串)"""
    if gb is None or pd.isna(gb):
        return ""
    try:
        num = float(gb)
        if num <= 0:
            return ""
        if num >= 1024:
            tb = num / 1024.0
            tb_str = f"{tb:.0f}" if tb.is_integer() else f"{tb:.1f}"
            return f"{tb_str} TB"
        num_str = f"{int(num)}" if num.is_integer() else f"{num:.1f}"
        return f"{num_str} GB"
    except (ValueError, TypeError):
        return f"{gb} GB" if gb else ""


def get_all_host_ports_list(h: Host) -> List[str]:
    """汇总主机自身与所有关联服务的开放端口 (去重排序)"""
    ports = []
    seen = set()

    def add_ports(raw_str: str):
        if not raw_str:
            return
        for p in re.split(r"[,，\s\n]+", str(raw_str).strip()):
            p = p.strip()
            if p and p not in seen:
                seen.add(p)
                ports.append(p)

    add_ports(h.open_ports)
    if hasattr(h, "cluster_relations") and h.cluster_relations:
        for rel in h.cluster_relations:
            if rel.cluster and rel.cluster.port:
                add_ports(rel.cluster.port)
    return ports


def clean_kernel_version(kernel: Any) -> str:
    """与主机资产列表完全一致的内核版本格式化 (仅提取 x.xx.x 主版本号，如 5.15.0-89-generic -> 5.15.0)"""
    if not kernel:
        return ""
    s = str(kernel).strip()
    if not s or s == "-":
        return ""
    m = re.match(r"^(\d+(\.\d+)+)", s)
    if m:
        return m.group(1)
    s_clean = re.sub(r"(\.el\d+.*?|\.x86_64|\.aarch64)", "", s)
    return s_clean.split("-")[0].strip()


def export_hosts_data(
    hosts: List[Host],
    file_format: str = "xlsx",
    selected_columns: Optional[List[str]] = None,
    meta_config: Optional[Dict[str, Any]] = None
) -> Tuple[io.BytesIO, str, str]:
    """将主机列表导出为 Excel 或 CSV 二进制流 (支持动态自定义列与字典转换，空数据纯空白展示)"""
    meta_config = meta_config or {}
    env_list = meta_config.get("environments", [])
    status_list = meta_config.get("host_statuses") or meta_config.get("statuses") or []
    arch_list = meta_config.get("cpu_architectures") or meta_config.get("architectures") or []

    # 字典转换映射
    env_map = {e.get("key"): e.get("label") for e in env_list if isinstance(e, dict) and e.get("key")}
    status_map = {s.get("key"): s.get("label") for s in status_list if isinstance(s, dict) and s.get("key")}
    arch_map = {a.get("key"): a.get("label") for a in arch_list if isinstance(a, dict) and a.get("key")}

    # 默认兜底映射
    default_env_fallback = {"prod": "生产环境", "test": "测试环境", "dev": "开发环境", "stage": "预发布环境"}
    default_status_fallback = {
        "online": "在线",
        "running": "在线",
        "offline": "下线",
        "stopped": "下线",
        "maintenance": "维护中",
        "warning": "告警",
        "error": "故障"
    }

    # 所有可用列定义 (key -> 中文标签)
    ALL_COLUMN_DEFS = [
        ("index", "序号"),
        ("hostname", "主机名"),
        ("private_ip", "内网IP"),
        ("public_ip", "外网IP"),
        ("open_ports", "开放端口"),
        ("env", "环境"),
        ("status", "状态"),
        ("cpu_cores", "CPU"),
        ("memory_gb", "内存"),
        ("disk_gb", "数据盘"),
        ("arch", "架构"),
        ("os", "操作系统"),
        ("kernel_version", "内核版本"),
        ("clusters", "所属服务"),
        ("notes", "备注"),
        ("created_at", "添加时间"),
        ("updated_at", "修改时间"),
    ]
    key_to_label = dict(ALL_COLUMN_DEFS)
    label_to_key = {label: key for key, label in ALL_COLUMN_DEFS}

    # 确定导出的字段序列
    if selected_columns and len(selected_columns) > 0:
        export_keys = []
        for col in selected_columns:
            col_str = str(col).strip()
            if col_str in key_to_label:
                export_keys.append(col_str)
            elif col_str in label_to_key:
                export_keys.append(label_to_key[col_str])
        # 如果解析后没有有效字段，则回退为全部
        if not export_keys:
            export_keys = [k for k, _ in ALL_COLUMN_DEFS]
    else:
        export_keys = [k for k, _ in ALL_COLUMN_DEFS]

    data = []
    for idx, h in enumerate(hosts, 1):
        # 1. 环境标签
        env_label = env_map.get(h.env) or default_env_fallback.get(h.env, h.env or "")

        # 2. 状态标签
        status_label = status_map.get(h.status) or default_status_fallback.get(h.status, h.status or "")

        # 3. 架构标签
        arch_label = arch_map.get(h.arch, h.arch or "")

        # 4. 汇总端口
        all_ports = get_all_host_ports_list(h)
        ports_display = ", ".join(all_ports) if all_ports else (h.open_ports or "")

        # 5. 拼接部署服务信息 (与主机资产列表徽章完全一致: 如 K8s v1.19.11, Redis v7.0.12)
        cluster_strs = []
        if hasattr(h, "cluster_relations") and h.cluster_relations:
            for rel in h.cluster_relations:
                if rel.cluster:
                    c_type = rel.cluster.cluster_type or rel.cluster.name or "服务"
                    raw_v = (rel.cluster.version or "").strip()
                    if raw_v:
                        c_ver = raw_v if (raw_v.startswith("v") or raw_v.startswith("V")) else f"v{raw_v}"
                        cluster_strs.append(f"{c_type} {c_ver}")
                    else:
                        cluster_strs.append(c_type)
                else:
                    cluster_strs.append(f"服务-{rel.cluster_id}")
        clusters_display = ", ".join(cluster_strs) if cluster_strs else ""

        # 6. 构造完整字段字典 (无数据时纯空白，无 - 占位符)
        cpu_display = f"{h.cpu_cores} 核" if (h.cpu_cores is not None and h.cpu_cores > 0) else ""
        row_dict = {
            "index": idx,
            "hostname": h.hostname or "",
            "private_ip": h.private_ip or "",
            "public_ip": h.public_ip or "",
            "open_ports": ports_display,
            "env": env_label,
            "status": status_label,
            "cpu_cores": cpu_display,
            "memory_gb": format_storage_str(h.memory_gb),
            "disk_gb": format_storage_str(h.disk_gb),
            "arch": arch_label,
            "os": h.os or "",
            "kernel_version": clean_kernel_version(h.kernel_version),
            "clusters": clusters_display,
            "notes": h.notes or "",
            "created_at": h.created_at.strftime("%Y-%m-%d %H:%M:%S") if h.created_at else "",
            "updated_at": h.updated_at.strftime("%Y-%m-%d %H:%M:%S") if h.updated_at else (h.created_at.strftime("%Y-%m-%d %H:%M:%S") if h.created_at else "")
        }

        # 根据选定的列输出
        filtered_row = {key_to_label[k]: row_dict.get(k, "") for k in export_keys}
        data.append(filtered_row)

    headers = [key_to_label[k] for k in export_keys]
    df = pd.DataFrame(data, columns=headers) if data else pd.DataFrame(columns=headers)

    output = io.BytesIO()
    if file_format.lower() == "csv":
        # CSV 使用 utf-8-sig 保证 Excel 打开不乱码
        df.to_csv(output, index=False, encoding="utf-8-sig")
        content_type = "text/csv; charset=utf-8-sig"
        extension = "csv"
    else:
        # Excel 格式化美化导出 (科技石板深蓝大厂高颜值方案)
        wb = Workbook()
        ws = wb.active
        ws.title = "主机资产清单"

        # 1. 科技石板深蓝 (Slate-800) 表头样式 (精致紧凑版)
        header_font = Font(name="Microsoft YaHei", size=10, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        header_border = Border(
            left=Side(style="thin", color="334155"),
            right=Side(style="thin", color="334155"),
            top=Side(style="thin", color="334155"),
            bottom=Side(style="medium", color="0EA5E9") # 底边青蓝强调线
        )
        thin_border = Border(
            left=Side(style="thin", color="E2E8F0"),
            right=Side(style="thin", color="E2E8F0"),
            top=Side(style="thin", color="E2E8F0"),
            bottom=Side(style="thin", color="E2E8F0")
        )
        alt_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
        white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

        export_headers = list(df.columns) if not df.empty else headers
        ws.append(export_headers)
        ws.row_dimensions[1].height = 22

        for col_idx in range(1, len(export_headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = header_border

        # 2. 数据字体与对齐规范 (精致小巧版)
        data_font = Font(name="Microsoft YaHei", size=9.5, color="0F172A")
        index_font = Font(name="Microsoft YaHei", size=9.5, color="64748B", bold=True)
        ip_font = Font(name="Consolas", size=9.5, color="2563EB", bold=True)
        center_align = Alignment(horizontal="center", vertical="center")
        left_align = Alignment(horizontal="left", vertical="center")

        # 状态徽章色彩映射 (9.5pt)
        status_styles = {
            "在线": (PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid"), Font(name="Microsoft YaHei", size=9.5, color="15803D", bold=True)),
            "online": (PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid"), Font(name="Microsoft YaHei", size=9.5, color="15803D", bold=True)),
            "维护中": (PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"), Font(name="Microsoft YaHei", size=9.5, color="B45309", bold=True)),
            "maintenance": (PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"), Font(name="Microsoft YaHei", size=9.5, color="B45309", bold=True)),
            "下线": (PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"), Font(name="Microsoft YaHei", size=9.5, color="B91C1C", bold=True)),
            "offline": (PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"), Font(name="Microsoft YaHei", size=9.5, color="B91C1C", bold=True)),
            "告警": (PatternFill(start_color="FFEDD5", end_color="FFEDD5", fill_type="solid"), Font(name="Microsoft YaHei", size=9.5, color="C2410C", bold=True)),
            "故障": (PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"), Font(name="Microsoft YaHei", size=9.5, color="B91C1C", bold=True)),
        }

        # 环境徽章色彩映射 (9.5pt)
        env_styles = {
            "生产环境": (PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid"), Font(name="Microsoft YaHei", size=9.5, color="1D4ED8", bold=True)),
            "生产": (PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid"), Font(name="Microsoft YaHei", size=9.5, color="1D4ED8", bold=True)),
            "prod": (PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid"), Font(name="Microsoft YaHei", size=9.5, color="1D4ED8", bold=True)),
            "测试环境": (PatternFill(start_color="F3E8FF", end_color="F3E8FF", fill_type="solid"), Font(name="Microsoft YaHei", size=9.5, color="7E22CE", bold=True)),
            "测试": (PatternFill(start_color="F3E8FF", end_color="F3E8FF", fill_type="solid"), Font(name="Microsoft YaHei", size=9.5, color="7E22CE", bold=True)),
            "test": (PatternFill(start_color="F3E8FF", end_color="F3E8FF", fill_type="solid"), Font(name="Microsoft YaHei", size=9.5, color="7E22CE", bold=True)),
            "开发环境": (PatternFill(start_color="E0F2FE", end_color="E0F2FE", fill_type="solid"), Font(name="Microsoft YaHei", size=9.5, color="0369A1", bold=True)),
            "开发": (PatternFill(start_color="E0F2FE", end_color="E0F2FE", fill_type="solid"), Font(name="Microsoft YaHei", size=9.5, color="0369A1", bold=True)),
            "dev": (PatternFill(start_color="E0F2FE", end_color="E0F2FE", fill_type="solid"), Font(name="Microsoft YaHei", size=9.5, color="0369A1", bold=True)),
            "预发布环境": (PatternFill(start_color="FCE7F3", end_color="FCE7F3", fill_type="solid"), Font(name="Microsoft YaHei", size=9.5, color="BE185D", bold=True)),
            "预发布": (PatternFill(start_color="FCE7F3", end_color="FCE7F3", fill_type="solid"), Font(name="Microsoft YaHei", size=9.5, color="BE185D", bold=True)),
            "stage": (PatternFill(start_color="FCE7F3", end_color="FCE7F3", fill_type="solid"), Font(name="Microsoft YaHei", size=9.5, color="BE185D", bold=True)),
        }

        center_cols = {"序号", "环境", "状态", "CPU", "内存", "数据盘", "架构", "内核版本", "添加时间", "修改时间"}

        for row_idx, row in df.iterrows():
            row_data = list(row.values)
            ws.append(row_data)
            current_row = ws.max_row
            ws.row_dimensions[current_row].height = 20
            is_alt = (row_idx % 2 == 1)

            for col_idx in range(1, len(row_data) + 1):
                cell = ws.cell(row=current_row, column=col_idx)
                cell.font = data_font
                cell.border = thin_border
                cell.fill = alt_fill if is_alt else white_fill

                col_name = export_headers[col_idx - 1]
                val_str = str(cell.value or "").strip()

                if col_name in center_cols:
                    cell.alignment = center_align
                else:
                    cell.alignment = left_align

                # 序号列样式
                if col_name == "序号":
                    cell.font = index_font

                # IP 列样式
                elif col_name in ["内网IP", "外网IP"] and val_str and val_str != "-":
                    cell.font = ip_font

                # 状态彩色徽章
                elif col_name == "状态" and val_str in status_styles:
                    s_fill, s_font = status_styles[val_str]
                    cell.fill = s_fill
                    cell.font = s_font

                # 环境彩色徽章
                elif col_name == "环境" and val_str in env_styles:
                    e_fill, e_font = env_styles[val_str]
                    cell.fill = e_fill
                    cell.font = e_font

        # 3. 智能跨行合并连续相同的外网IP单元格 (支持多台主机共享同一外网IP的合并展示)
        if "外网IP" in export_headers:
            pub_col_idx = export_headers.index("外网IP") + 1
            start_r = 2
            max_r = ws.max_row
            while start_r <= max_r:
                val = str(ws.cell(row=start_r, column=pub_col_idx).value or "").strip()
                if not val or val == "-":
                    start_r += 1
                    continue
                end_r = start_r
                while end_r + 1 <= max_r:
                    next_val = str(ws.cell(row=end_r + 1, column=pub_col_idx).value or "").strip()
                    if next_val == val:
                        end_r += 1
                    else:
                        break
                if end_r > start_r:
                    ws.merge_cells(
                        start_row=start_r,
                        start_column=pub_col_idx,
                        end_row=end_r,
                        end_column=pub_col_idx
                    )
                    # 重新应用边框和垂直居中对齐，确保合并单元格样式美观完整
                    for r in range(start_r, end_r + 1):
                        ws.cell(row=r, column=pub_col_idx).border = thin_border
                    ws.cell(row=start_r, column=pub_col_idx).alignment = Alignment(horizontal="left", vertical="center")
                    start_r = end_r + 1
                else:
                    start_r += 1

        # 4. 智能计算紧凑列宽 (兼顾内容宽度与适度留白，拒绝过宽臃肿)
        min_col_widths = {
            "序号": 6,
            "环境": 9,
            "状态": 9,
            "CPU": 7,
            "内存": 9,
            "数据盘": 9,
            "架构": 9,
            "内网IP": 14,
            "外网IP": 14,
            "开放端口": 13,
            "主机名": 16,
            "操作系统": 16,
            "内核版本": 15,
            "所属服务": 18,
            "备注": 14,
            "添加时间": 16,
            "修改时间": 16,
        }

        for col_idx, col in enumerate(ws.columns, 1):
            col_name = export_headers[col_idx - 1] if col_idx - 1 < len(export_headers) else ""
            default_min = min_col_widths.get(col_name, 10)
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val = str(cell.value or "")
                try:
                    val_len = len(val.encode('gbk', 'ignore'))
                except Exception:
                    val_len = len(val)
                max_len = max(max_len, val_len)
            ws.column_dimensions[col_letter].width = max(max_len + 3, default_min)

        # 5. 视图交互增强: 开启筛选器、冻结首行、显示网格线
        ws.freeze_panes = "A2"
        if ws.max_row > 1 and export_headers:
            ws.auto_filter.ref = f"A1:{get_column_letter(len(export_headers))}{ws.max_row}"
        if ws.views.sheetView:
            ws.views.sheetView[0].showGridLines = True

        wb.save(output)
        content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        extension = "xlsx"

    output.seek(0)
    return output, content_type, extension
